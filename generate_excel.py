import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
DATA_FONT = Font(name="Calibri", size=10)
REMOVE_FONT = Font(name="Calibri", size=10, color="9C0006")
REMOVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
)

tables_data = {
    "Overview": {
        "title": "TỔNG HỢP CÁC ĐIỂM ĐÃ TỐI ƯU & LƯỢC BỎ TRONG CSDL (NUTRISMART AGENT)",
        "headers": ["STT", "Tên Bảng / Thuộc tính", "Nhóm Chức Năng", "Trạng Thái", "Đánh Giá & Lý Do Lược Bỏ"],
        "rows": [
            [1, "role_permissions", "Người dùng & Phân quyền", "✅ THAY THẾ", "Quyền dùng chung theo role, thay cho quyền riêng theo từng staff user_id"],
            [2, "nutrition_plans.parent_plan_id", "Kế hoạch Dinh dưỡng", "❌ ĐÃ LƯỢC BỎ", "Dư thừa vì đã có user_id + version + created_at để xác định thứ tự"],
            [3, "body_metrics_history.bmi", "Người dùng & Hồ sơ", "❌ ĐÃ LƯỢC BỎ", "Dư thừa vì BMI được tính động từ height_cm và weight_kg tại từng mốc"],
            [4, "documents.raw_text", "Tri thức RAG", "❌ ĐÃ LƯỢC BỎ", "Dư thừa vì RAG chỉ search theo doc_chunks, raw_text làm nặng DB"],
            [5, "users", "Người dùng & Hồ sơ", "✅ GIỮ NGUYÊN", "Tài khoản người dùng (email, password_hash, role)"],
            [6, "activity_levels", "Người dùng & Hồ sơ", "✅ BỔ SUNG", "Danh mục mức vận động và hệ số tính TDEE; không cần thuộc tính code"],
            [7, "medical_conditions", "Người dùng & Hồ sơ", "✅ GIỮ NGUYÊN", "Danh mục bệnh nền"],
            [8, "allergens", "Người dùng & Hồ sơ", "✅ GIỮ NGUYÊN", "Danh mục chất gây dị ứng"],
            [9, "user_medical_conditions", "Người dùng & Hồ sơ", "✅ ĐỔI TÊN", "Bảng liên kết Người dùng - Bệnh nền"],
            [10, "user_allergens", "Người dùng & Hồ sơ", "✅ ĐỔI TÊN", "Bảng liên kết Người dùng - Dị ứng; đã bỏ severity không dùng"],
            [11, "body_metrics_history", "Người dùng & Hồ sơ", "✅ GIỮ NGUYÊN", "Nhật ký chiều cao và cân nặng theo ngày"],
            [12, "foods", "Nhật ký & Vận động", "✅ GIỮ NGUYÊN", "Thực phẩm (giữ cột source để biết nguồn AI/Người dùng/USDA)"],
            [13, "meal_images", "Nhật ký & Vận động", "✅ GIỮ NGUYÊN", "Ảnh bữa ăn phân tích bởi AI Vision"],
            [14, "meal_logs", "Nhật ký & Vận động", "✅ GIỮ NGUYÊN", "Nhật ký khẩu phần ăn thực tế"],
            [15, "exercises", "Nhật ký & Vận động", "✅ GIỮ NGUYÊN", "Thư viện bài tập vận động"],
            [16, "activity_logs", "Nhật ký & Vận động", "✅ GIỮ NGUYÊN", "Nhật ký vận động / đếm bước chân theo ngày"],
            [17, "nutrition_plans", "Kế hoạch Dinh dưỡng", "✅ GIỮ NGUYÊN", "Lộ trình thực đơn AI/Chuyên gia"],
            [18, "plan_evaluations", "Kế hoạch Dinh dưỡng", "✅ GIỮ NGUYÊN", "Đánh giá lộ trình chu kỳ 7 ngày"],
            [19, "doc_categories", "Tri thức RAG", "✅ GIỮ NGUYÊN", "Phân loại tài liệu"],
            [20, "documents", "Tri thức RAG", "✅ GIỮ NGUYÊN", "Tài liệu y tế gốc (giữ file_path)"],
            [21, "doc_chunks", "Tri thức RAG", "✅ GIỮ NGUYÊN", "Đoạn văn bản cắt nhỏ + Vector Embedding"],
            [22, "drug_categories", "Dược phẩm", "✅ GIỮ NGUYÊN", "Phân loại nhóm thuốc"],
            [23, "drugs", "Dược phẩm", "✅ GIỮ NGUYÊN", "Chi tiết biệt dược, chỉ định, chống chỉ định"],
            [24, "chat_sessions", "Chat AI & System", "✅ GIỮ NGUYÊN", "Phiên trò chuyện với AI"],
            [25, "chat_messages", "Chat AI & System", "✅ GIỮ NGUYÊN", "Nội dung tin nhắn người dùng & AI"],
            [26, "message_citations", "Chat AI & System", "✅ GIỮ NGUYÊN", "Trích dẫn nguồn RAG"],
            [27, "notifications", "Chat AI & System", "✅ GIỮ NGUYÊN", "Thông báo người dùng"],
            [28, "audit_logs", "Chat AI & System", "✅ GIỮ NGUYÊN", "Nhật ký thao tác quản trị"]
        ]
    },
    "users": {
        "headers": ["id", "email", "password_hash", "full_name", "role", "created_at", "updated_at"],
        "rows": [
            [1, "user1@nutrismart.vn", "$2b$12$eImiTXuWV...", "Nguyễn Văn An", "USER", "2026-07-01 08:00:00", "2026-07-01 08:00:00"],
            [2, "expert1@nutrismart.vn", "$2b$12$8xKk12aL...", "BS. Trần Thu Hà", "EXPERT", "2026-06-15 09:30:00", "2026-07-10 10:15:00"],
            [3, "admin@nutrismart.vn", "$2b$12$9yMm34bM...", "Quản Trị Viên", "ADMIN", "2026-06-01 00:00:00", "2026-06-01 00:00:00"]
        ]
    },
    "user_profile": {
        "headers": ["user_id", "gender", "birth_date", "activity_level_id", "goal", "daily_calorie_target", "updated_at"],
        "rows": [
            [1, "MALE", "1995-04-12", 2, "LOSE_WEIGHT", 1850, "2026-07-20 14:00:00"],
            [2, "FEMALE", "1990-11-20", 4, "MAINTAIN", 1900, "2026-07-15 16:30:00"]
        ]
    },
    "activity_levels": {
        "headers": ["id", "name", "description", "calorie_multiplier"],
        "rows": [
            [1, "Ít vận động", "Hầu như không tập thể dục", 1.200],
            [2, "Vận động nhẹ", "Tập nhẹ 1-3 ngày/tuần", 1.375],
            [3, "Vận động vừa", "Tập vừa 3-5 ngày/tuần", 1.550],
            [4, "Vận động nhiều", "Tập nặng 6-7 ngày/tuần", 1.725],
            [5, "Vận động rất nhiều", "Tập rất nặng hoặc lao động thể lực", 1.900]
        ]
    },
    "role_permissions": {
        "headers": ["role", "can_manage_users", "can_manage_foods", "can_manage_categories", "can_review_documents", "can_review_plans", "can_review_ai_chat", "can_review_logs", "can_manage_permissions"],
        "rows": [
            ["USER", False, False, False, False, False, False, False, False],
            ["EXPERT", False, False, False, True, True, True, True, False],
            ["ADMIN", True, True, True, True, True, True, True, True]
        ]
    },
    "medical_conditions": {
        "headers": ["id", "name"],
        "rows": [
            [1, "Đái tháo đường týp 2"],
            [2, "Tăng huyết áp vô căn"],
            [3, "Trào ngược dạ dày thực quản (GERD)"]
        ]
    },
    "allergens": {
        "headers": ["id", "name"],
        "rows": [
            [1, "Hải sản (Tôm, Cua, Mực)"],
            [2, "Đậu phộng / Lạc"],
            [3, "Sữa & Lactose"]
        ]
    },
    "user_medical_conditions": {
        "headers": ["user_id", "condition_id"],
        "rows": [
            [1, 1],
            [1, 2]
        ]
    },
    "user_allergens": {
        "headers": ["user_id", "allergen_id"],
        "rows": [
            [1, 1]
        ]
    },
    "body_metrics_history": {
        "headers": ["id", "user_id", "recorded_at", "height_cm", "weight_kg"],
        "rows": [
            [1, 1, "2026-07-20", 170.00, 75.00],
            [2, 1, "2026-07-27", 170.00, 74.20]
        ]
    },
    "foods": {
        "headers": ["id", "name", "serving_desc", "serving_gram", "calories_kcal", "protein_g", "carb_g", "fat_g", "source"],
        "rows": [
            [1, "Phở bò tái", "1 tô vừa", 450.00, 480.00, 24.50, 58.00, 15.00, "USDA"],
            [2, "Ức gà áp chảo", "1 miếng (150g)", 150.00, 247.50, 46.50, 0.00, 5.40, "AI Gemini Flash"]
        ]
    },
    "meal_images": {
        "headers": ["id", "user_id", "image_path", "status", "predicted_food_id", "confidence", "raw_prediction", "estimated_kcal", "suitability_note", "created_at"],
        "rows": [
            [1, 1, "/uploads/meals/20260731_lunch.jpg", "DONE", 1, 0.925, '{"dish": "Phở bò", "ingredients": ["bánh phở", "thịt bò"]}', 480.00, "Phù hợp bữa trưa, nhưng chứa nhiều natri.", "2026-07-31 12:15:00"]
        ]
    },
    "meal_logs": {
        "headers": ["id", "user_id", "food_id", "meal_image_id", "meal_type", "quantity", "calories_kcal", "logged_at", "log_date"],
        "rows": [
            [1, 1, 1, 1, "LUNCH", 1.00, 480.00, "2026-07-31 12:20:00", "2026-07-31"]
        ]
    },
    "exercises": {
        "headers": ["id", "name", "met_value", "category"],
        "rows": [
            [1, "Chạy bộ (8 km/h)", 8.00, "Cardio"],
            [2, "Gập bụng / Gym", 3.80, "Kháng lực"]
        ]
    },
    "activity_logs": {
        "headers": ["id", "user_id", "exercise_id", "steps", "duration_min", "calories_burned", "log_date"],
        "rows": [
            [1, 1, 1, 0, 30, 280.00, "2026-07-31"],
            [2, 1, None, 8500, 60, 240.00, "2026-07-31"]
        ]
    },
    "nutrition_plans": {
        "headers": ["id", "user_id", "version", "start_date", "end_date", "daily_kcal_target", "goal", "content", "status", "created_at"],
        "rows": [
            [1, 1, 1, "2026-07-20", "2026-07-26", 1850, "LOSE_WEIGHT", '{"day_1": {"breakfast": "Phở gà ít bánh", "lunch": "Cơm lứt + ức gà"}}', "ACTIVE", "2026-07-19 20:00:00"]
        ]
    },
    "plan_evaluations": {
        "headers": ["id", "plan_id", "period_start", "period_end", "avg_kcal_intake", "weight_change_kg", "result", "ai_feedback", "evaluated_at"],
        "rows": [
            [1, 1, "2026-07-20", "2026-07-26", 1820.50, -0.80, "ACHIEVED", "Người dùng tuân thủ tốt mục tiêu calo, cân nặng giảm đúng tiến độ.", "2026-07-27 01:00:00"]
        ]
    },
    "doc_categories": {
        "headers": ["id", "parent_id", "name", "slug"],
        "rows": [
            [1, None, "Hướng dẫn Dinh dưỡng Quốc gia", "huong-dan-dinh-duong"],
            [2, 1, "Dinh dưỡng cho bệnh tiểu đường", "dinh-duong-tieu-duong"]
        ]
    },
    "documents": {
        "headers": ["id", "category_id", "title", "source_url", "source_name", "language", "file_path", "status", "uploaded_by", "approved_by", "approved_at", "created_at", "deleted_at"],
        "rows": [
            [1, 2, "Hướng dẫn chế độ ăn cho người tiểu đường týp 2", "https://moh.gov.vn/guidelines/diabetes.pdf", "Bộ Y Tế Việt Nam", "vi", "/storage/docs/diabetes.pdf", "APPROVED", 2, 3, "2026-06-20 10:00:00", "2026-06-18 15:30:00", None]
        ]
    },
    "doc_chunks": {
        "headers": ["id", "document_id", "chunk_index", "content", "token_count", "embedding", "metadata"],
        "rows": [
            [1, 1, 0, "Người bệnh tiểu đường nên ưu tiên thực phẩm có chỉ số đường huyết thấp...", 128, "[0.0123, -0.0456, 0.0891, ...]", '{"page": 3, "section": "Chỉ số GI"}']
        ]
    },
    "drug_categories": {
        "headers": ["id", "name"],
        "rows": [
            [1, "Thuốc điều trị tiểu đường"],
            [2, "Thuốc hạ huyết áp"]
        ]
    },
    "drugs": {
        "headers": ["id", "category_id", "document_id", "name", "active_ingredient", "indications", "side_effects", "contraindications", "status", "status_note", "created_at", "deleted_at"],
        "rows": [
            [1, 1, 1, "Glucophage 850mg", "Metformin Hydrochloride", "Điều trị bệnh tiểu đường týp 2", "Rối loạn tiêu hóa, tiêu chảy", "Suy thận nặng, nhiễm toan chuyển hóa", "ALLOWED", "Sử dụng theo đơn bác sĩ", "2026-06-21 09:00:00", None]
        ]
    },
    "chat_sessions": {
        "headers": ["id", "user_id", "title", "created_at"],
        "rows": [
            [1, 1, "Tư vấn thực đơn cho người tiểu đường", "2026-07-31 14:00:00"]
        ]
    },
    "chat_messages": {
        "headers": ["id", "session_id", "role", "content", "flagged", "created_at"],
        "rows": [
            [1, 1, "user", "Tôi bị tiểu đường týp 2 thì sáng nên ăn gì?", False, "2026-07-31 14:00:01"],
            [2, 1, "assistant", "Chào bạn, bạn nên ăn các món có chỉ số GI thấp như bún tươi, phở ít bánh nhiều rau...", False, "2026-07-31 14:00:05"]
        ]
    },
    "message_citations": {
        "headers": ["message_id", "chunk_id", "score", "rank"],
        "rows": [
            [2, 1, 0.8950, 1]
        ]
    },
    "notifications": {
        "headers": ["id", "user_id", "type", "title", "body", "is_read", "created_at"],
        "rows": [
            [1, 1, "PLAN_EVAL", "Đánh giá lộ trình tuần", "Lộ trình tuần qua của bạn đã hoàn thành xuất sắc!", False, "2026-07-27 01:05:00"]
        ]
    },
    "audit_logs": {
        "headers": ["id", "actor_id", "action", "entity", "entity_id", "before_data", "after_data", "ip_address", "created_at"],
        "rows": [
            [1, 3, "APPROVE_DOCUMENT", "documents", "1", '{"status": "PENDING"}', '{"status": "APPROVED"}', "192.168.1.50", "2026-06-20 10:00:00"]
        ]
    }
}

for sheet_name, data in tables_data.items():
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True
    
    start_row = 1
    if "title" in data:
        ws.cell(row=1, column=1, value=data["title"]).font = TITLE_FONT
        start_row = 3
    
    headers = data["headers"]
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header_title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    
    for row_idx, row_values in enumerate(data["rows"], start_row + 1):
        is_removed = "❌" in str(row_values[3]) if len(row_values) > 3 else False
        for col_num, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=val)
            cell.font = REMOVE_FONT if is_removed else DATA_FONT
            if is_removed:
                cell.fill = REMOVE_FILL
            cell.border = THIN_BORDER
            
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif isinstance(val, bool):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < start_row:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 65)

output_path = "NutriSmart_Database_Sample_Data_Optimized.xlsx"
wb.save(output_path)
print(f"Successfully updated: {output_path}")
